# -*- coding: UTF-8 -*-
#! python3

"""
    Usage from the repo root folder:

    ```python
    python -m migrate_from_excel
    ```
"""

# #############################################################################
# ########## Libraries #############
# ##################################

# Standard library
from os import environ
from pathlib import Path
from time import sleep
from timeit import default_timer
import logging
import urllib3

# 3rd party
from dotenv import load_dotenv

from openpyxl import load_workbook

# module target
from isogeo_pysdk import (
    IsogeoSession,
    __version__ as pysdk_version,
    Catalog,
    Contact,
    License,
    Metadata,
    Specification,
    Workgroup,
)

from isogeo_pysdk import IsogeoChecker

# #############################################################################
# ######## Globals #################
# ##################################

checker = IsogeoChecker()

# logs
logger = logging.getLogger(__name__)
logging.captureWarnings(True)
logger.setLevel(logging.DEBUG)

# paths
path_input_excel = Path(Path(__file__).parent, "_isogeo_complet_ready_migration.xlsx")

# -- SOURCE PARAMS
group = ""
platform = "qa"
load_dotenv("{}.env".format(platform), override=True)  # environment vars

# handle warnings
if platform.lower() == "qa":
    urllib3.disable_warnings()

# #############################################################################
# ########## Main program ###############
# #######################################

START_TIME = default_timer()  # chrono

# -- LOAD Excel -----
# print(path_input_excel.resolve())
wb = load_workbook(filename=path_input_excel.resolve(), read_only=True)
ws_contacts = wb["Contacts"]
ws_vectors = wb["Vecteurs"]

print("Excel workbook loaded at {:5.2f}s".format(default_timer() - START_TIME))

# -- AUTH --------
# Isogeo client
isogeo = IsogeoSession(
    client_id=environ.get("ISOGEO_API_USER_CLIENT_ID"),
    client_secret=environ.get("ISOGEO_API_USER_CLIENT_SECRET"),
    auto_refresh_url="{}/oauth/token".format(environ.get("ISOGEO_ID_URL")),
    platform=environ.get("ISOGEO_PLATFORM", "qa"),
)

# getting a token
isogeo.connect(
    username=environ.get("ISOGEO_USER_NAME"),
    password=environ.get("ISOGEO_USER_PASSWORD"),
)

print("Authentication succeeded at {:5.2f}s".format(default_timer() - START_TIME))

# -- Parse Excel workbook --
"""
    Isogeo to Office structure:

        Metadata UUID   = AN = 39
        Title           = A  = 0
        Technical name  = B  = 1
        Abstract        = C  = 2
        Keywords        = F  = 6
        INSPIRE         = G  = 7
        Context         = J  = 9
        Method          = K  = 10
        Frequency       = N  = 13
"""

dct_i2o_struct = {
    "uuid": 39,
    "title": 0,
    "name": 1,
    "abstract": 2,
    "keywords": 6,
    "inspire": 7,
    "context": 8,
    "method": 9,
    "frequency": 13,
    "date_creation": 14,
    "date_update": 16,
    "date_publication": 17,
    "scale": 23,
    "conditions": 30,
    "contacts": 32,
}


# for row in ws_vectors.rows:
# for cell in row:
# print(row[39].value) # metadata UUID
# print(row[0].value) # title
# print(row[2].value) # abstract
# print(row[5].value) # keywords
# print(row[6].value) # INSPIRE
# pass

for row in ws_vectors.iter_rows(min_row=2):
    logger.info("Reading metadata row: " + row[dct_i2o_struct.get("uuid")].value)

    # get metadata UUID from workbook
    metadata_uuid = row[dct_i2o_struct.get("uuid")].value
    if row[dct_i2o_struct.get("uuid")] is not None and checker.check_is_uuid(
        metadata_uuid
    ):
        metadata_uuid = row[dct_i2o_struct.get("uuid")].value
    else:
        logger.error("Invalid metadata UUID spotted: " + metadata_uuid)
        continue

    # compare title and technical names
    if row[dct_i2o_struct.get("title")].value == row[dct_i2o_struct.get("name")].value:
        logger.warning("Row has been ignored because title has not been changed.")
        continue
    else:
        pass

    # retrieve the metadata from Isogeo
    md_dict = isogeo.resource(resource_id=metadata_uuid)
    target_md = Metadata(**md_dict)
    # print(md_dict.get("name"), target_md.name)
    # print(target_md.name == row[dct_i2o_struct.get("name")].value)
    # check if technical names are matching
    if target_md.name == row[dct_i2o_struct.get("name")].value:
        logger.info("This a technical match! Show can go on!")
    else:
        logger.error("Hmmmm, there is no match between technical names")
        continue

    # contacts
    # print(row[dct_i2o_struct.get("contacts")].value)
    contacts_uuids = (
        "f628a23c260b46cea83c98f6c1655119",
        "baf146d7befa474b94f19f25b92915ea",
    )

    for contact in contacts_uuids:
        isogeo.md_associate_listing(target_md, contact)

    # events
    if row[dct_i2o_struct.get("date_creation")].value:
        try:
            isogeo.md_associate_events(
                metadata=target_md,
                event_date=row[dct_i2o_struct.get("date_creation")].value,
                event_kind="creation",
            )
        except Exception as e:
            logger.error(e)

    if row[dct_i2o_struct.get("date_update")].value:
        try:
            isogeo.md_associate_events(
                metadata=target_md,
                event_date=row[dct_i2o_struct.get("date_update")].value,
                event_kind="update",
            )
        except Exception as e:
            logger.error(e)

    if row[dct_i2o_struct.get("date_publication")].value:
        try:
            isogeo.md_associate_events(
                metadata=target_md,
                event_date=row[dct_i2o_struct.get("date_publication")].value,
                event_kind="publication",
            )
        except Exception as e:
            logger.error(e)

    # print(row[dct_i2o_struct.get("context")].value)

    # edit local metadata object
    target_md.title = row[dct_i2o_struct.get("title")].value
    target_md.abstract = row[dct_i2o_struct.get("abstract")].value
    target_md.collectionContext = row[dct_i2o_struct.get("context")].value
    target_md.collectionMethod = row[dct_i2o_struct.get("method")].value
    target_md.updateFrequency = row[dct_i2o_struct.get("frequency")].value
    target_md.scale = row[dct_i2o_struct.get("scale")].value
    # print(row[39].value) # metadata UUID
    # print(row[0].value) # title
    # print(row[2].value) # abstract
    # print(row[5].value) # keywords
    # print(row[6].value)  # INSPIRE

    # update online metadata
    isogeo.md_update(target_md)
    sleep(1)
    logger.info(
        "{} update finished at {:5.2f}s".format(
            metadata_uuid, default_timer() - START_TIME
        )
    )

# correctly close connection
isogeo.close()

print("Import finished at {:5.2f}s".format(default_timer() - START_TIME))

# ##############################################################################
# ##### Stand alone program ########
# ##################################
if __name__ == "__main__":
    pass
